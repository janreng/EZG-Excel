"""Đọc/ghi file CSV và XLSX.

XLSX giữ cả **định dạng** (phông, màu nền/chữ, viền, căn lề, định dạng số) và
**ô gộp** (merge). CSV chỉ lưu giá trị thô.

API:
  load_file(path) -> (rows, fmt, merges)
  save_file(path, rows, fmt=None, merges=None)
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _normalize(rows: list[list], min_cols: int = 1) -> list[list[str]]:
    """Chuẩn hóa thành ma trận chữ nhật toàn chuỗi."""
    rows = [[("" if v is None else str(v)) for v in row] for row in rows]
    width = max([len(r) for r in rows] + [min_cols])
    for r in rows:
        r.extend([""] * (width - len(r)))
    if not rows:
        rows = [[""] * width]
    return rows


# ---------------------------------------------------------------- CSV


def load_csv(path: str | Path) -> list[list[str]]:
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(f, dialect))
    return _normalize(rows)


def save_csv(path: str | Path, rows: list[list[str]]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)


# ---------------------------------------------------------------- XLSX: màu


def _to_argb(hex_color: str) -> str:
    """'#RRGGBB' -> 'FFRRGGBB' (openpyxl dùng aRGB)."""
    h = hex_color.lstrip("#").upper()
    if len(h) == 6:
        return "FF" + h
    return h


def _from_argb(color) -> str | None:
    """openpyxl Color -> '#RRGGBB', bỏ qua màu theo theme/None."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    return "#" + rgb[-6:]


# ---------------------------------------------------------------- XLSX: lưu


def _build_style(fmt: dict) -> dict:
    """Đổi dict định dạng của app -> các đối tượng style openpyxl."""
    style: dict = {}
    font_kw = {}
    if fmt.get("font"):
        font_kw["name"] = fmt["font"]
    if fmt.get("size"):
        font_kw["size"] = int(fmt["size"])
    if fmt.get("bold"):
        font_kw["bold"] = True
    if fmt.get("italic"):
        font_kw["italic"] = True
    if fmt.get("underline"):
        font_kw["underline"] = "single"
    if fmt.get("strike"):
        font_kw["strike"] = True
    if fmt.get("color"):
        font_kw["color"] = _to_argb(fmt["color"])
    if font_kw:
        style["font"] = Font(**font_kw)

    if fmt.get("bg"):
        style["fill"] = PatternFill(fill_type="solid", fgColor=_to_argb(fmt["bg"]))

    align_kw = {}
    if fmt.get("halign"):
        align_kw["horizontal"] = fmt["halign"]
    valign = fmt.get("valign")
    if valign:
        align_kw["vertical"] = "center" if valign == "middle" else valign
    if fmt.get("wrap") == "wrap":
        align_kw["wrap_text"] = True
    if align_kw:
        style["alignment"] = Alignment(**align_kw)

    border = fmt.get("border")
    if border:
        sides = {}
        for side, color in border.items():
            sides[side] = Side(style="thin", color=_to_argb(color))
        style["border"] = Border(**sides)

    if fmt.get("number_format"):
        style["number_format"] = fmt["number_format"]
    return style


def _write_ws(ws, rows: list[list[str]], fmt: dict, merges: list) -> None:
    """Ghi dữ liệu + định dạng + ô gộp vào một worksheet."""
    from .formula import is_formula

    fmt = fmt or {}
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            cell_fmt = fmt.get((r - 1, c - 1))
            if value == "" and not cell_fmt:
                continue
            cell = ws.cell(row=r, column=c)
            if value != "":
                if is_formula(value):
                    cell.value = value
                else:
                    num = _try_number(value)
                    cell.value = num if num is not None else value
            if cell_fmt:
                for attr, obj in _build_style(cell_fmt).items():
                    setattr(cell, attr, obj)

    for (t, l, b, r) in (merges or []):
        ws.merge_cells(
            start_row=t + 1, start_column=l + 1, end_row=b + 1, end_column=r + 1
        )


def save_xlsx(path: str | Path, sheets: list) -> None:
    """sheets: danh sách (name, rows, fmt, merges)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định, tự tạo theo danh sách
    for name, rows, fmt, merges in sheets:
        ws = wb.create_sheet(title=_safe_sheet_name(name))
        _write_ws(ws, rows, fmt, merges)
    if not wb.sheetnames:  # phòng trường hợp rỗng
        wb.create_sheet(title="Sheet1")
    wb.save(path)


def _safe_sheet_name(name: str) -> str:
    """Tên sheet hợp lệ cho Excel (<=31 ký tự, bỏ ký tự cấm)."""
    for ch in r'[]:*?/\\':
        name = name.replace(ch, " ")
    return (name.strip() or "Sheet")[:31]


def _try_number(text: str):
    try:
        if "." in text or "e" in text.lower():
            return float(text)
        return int(text)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------- XLSX: đọc


def _read_fmt(cell) -> dict:
    """Đọc định dạng của một ô openpyxl -> dict định dạng của app."""
    f: dict = {}
    font = cell.font
    if font is not None:
        if font.name and font.name != "Calibri":
            f["font"] = font.name
        if font.size and int(font.size) != 11:
            f["size"] = int(font.size)
        if font.bold:
            f["bold"] = True
        if font.italic:
            f["italic"] = True
        if font.underline:
            f["underline"] = True
        if font.strike:
            f["strike"] = True
        color = _from_argb(font.color)
        if color and color != "#000000":
            f["color"] = color

    fill = cell.fill
    if fill is not None and fill.fill_type == "solid":
        bg = _from_argb(fill.fgColor)
        if bg and bg != "#000000":
            f["bg"] = bg

    align = cell.alignment
    if align is not None:
        if align.horizontal in ("left", "center", "right"):
            f["halign"] = align.horizontal
        if align.vertical in ("top", "center", "bottom"):
            f["valign"] = "middle" if align.vertical == "center" else align.vertical
        if align.wrap_text:
            f["wrap"] = "wrap"

    border = cell.border
    if border is not None:
        b = {}
        for side in ("top", "bottom", "left", "right"):
            s = getattr(border, side, None)
            if s is not None and s.style:
                b[side] = _from_argb(s.color) or "#000000"
        if b:
            f["border"] = b

    nf = cell.number_format
    if nf and nf != "General":
        f["number_format"] = nf
    return f


def _read_ws(ws) -> tuple[list[list[str]], dict, list]:
    """Đọc một worksheet -> (rows, fmt, merges)."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    rows: list[list[str]] = []
    fmt: dict = {}
    for r in range(1, max_row + 1):
        row_vals: list[str] = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            row_vals.append("" if v is None else str(v))
            cell_fmt = _read_fmt(cell)
            if cell_fmt:
                fmt[(r - 1, c - 1)] = cell_fmt
        rows.append(row_vals)

    merges = [
        (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
        for rng in ws.merged_cells.ranges
    ]
    return _normalize(rows), fmt, merges


def load_xlsx(path: str | Path) -> list:
    """Đọc mọi sheet -> danh sách (name, rows, fmt, merges)."""
    # data_only=False để giữ nguyên công thức (=...) — app tự tính lại.
    # Không dùng read_only để đọc được đầy đủ style + merge.
    wb = openpyxl.load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        rows, fmt, merges = _read_ws(ws)
        sheets.append((ws.title, rows, fmt, merges))
    wb.close()
    return sheets or [("Sheet1", _normalize([[""]]), {}, [])]


# ---------------------------------------------------------------- điều phối


def load_file(path: str | Path) -> list:
    """Tự nhận định dạng theo phần mở rộng.

    Trả về danh sách sheet ``[(name, rows, fmt, merges), ...]``.
    CSV chỉ có một sheet (tên theo file).
    """
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_xlsx(path)
    if suffix in (".csv", ".txt", ".tsv"):
        return [(Path(path).stem or "Sheet1", load_csv(path), {}, [])]
    raise ValueError(f"Định dạng không hỗ trợ: {suffix}")


def save_file(path: str | Path, sheets: list) -> None:
    """sheets: danh sách (name, rows, fmt, merges).

    XLSX lưu mọi sheet kèm định dạng; CSV chỉ lưu giá trị của sheet đầu tiên.
    """
    suffix = Path(path).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        save_xlsx(path, sheets)
    elif suffix in (".csv", ".txt", ".tsv"):
        save_csv(path, sheets[0][1] if sheets else [[""]])  # CSV: chỉ sheet đầu
    else:
        raise ValueError(f"Định dạng không hỗ trợ: {suffix}")
